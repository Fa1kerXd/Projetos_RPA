"""
Modulo responsavel por preencher formulario de forma automatica capturando as informações
de uma planilha Excel.

Fluxo:
    1. iterar sobre as linhas da planilha pegando os valores.
    2. Preencher de forma automatizada os dados iterados em um formulario
    
"""

# Importando biblioteca selenium para a automatização do fluxo.
# Pyautogui para dar um intervalo entre processos.
from selenium import webdriver as driver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By 
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time 
from pathlib import Path



# Caminho especificado da PLanilha.
DIR = Path(__file__).parent
FILE = DIR / 'DadosFormulario.xlsx'
FORM_URL = "https://pt.surveymonkey.com/r/WLXYDX2"


FORM_FIELDS = {
    "nome": ("name", "166517069"),
    "email": ("name", "166517072"),
    "telefone": ("name", "166517070"),
    "genero_masculino": ("name", "question-field-166517071"),
    "genero_feminino": ("id", "1215509813"),
    "sobre": ("name", "166517073"),
    "submit": ("xpath", '//*[@id="view-pageNavigation"]/div/button')
}







def  wait_for_element(locator_type: str,value: str, browser: driver.Edge, timeout: int = 10):
    """
    Aguarda até que um elemento esteja presente na página.

    Args:
        browser (webdriver.Edge): Instância do navegador Selenium.
        locator_type (str): Tipo do localizador ('name', 'id', 'xpath', etc.).
        locator_value (str): Valor do localizador (ID, nome, XPath, etc.).
        timeout (int): Tempo máximo de espera em segundos (padrão: 10).

    Returns:
        WebElement: Elemento encontrado na página.

    Raises:
        TimeoutException: Se o elemento não for encontrado dentro do timeout.
    """
    wait = WebDriverWait(browser, timeout)

    locator_map = {
        "name": By.NAME,
        "id": By.ID,
        "xpath": By.XPATH,
        "class": By.CLASS_NAME
        }
    locator = locator_map.get(locator_type.lower())
    
    if not locator:
        raise ValueError(f"Tipo de localizador inválido: {locator_type}")    
        
    return  wait.until(EC.presence_of_element_located((locator, value)))



def preencher_planilha(linha_dados: dict) -> bool:
    """
    Preenche e submete o formulário web com os dados fornecidos.

    Args:
        linha_dados (dict): Dicionário contendo os dados da linha:
            - 'nome' (str): Nome completo
            - 'email' (str): Email
            - 'telefone' (str): Telefone
            - 'genero' (str): 'Masculino' ou 'Feminino'
            - 'sobre' (str): Descrição sobre a pessoa

    Returns:
        bool: True se o formulário foi preenchido e enviado com sucesso, False caso contrário.
    """

    browser = None

    try:
        # Inicializa o navegador
        browser = driver.Edge()
        browser.get(FORM_URL)
        print(f"Processando: {linha_dados['nome']}")



        # Preenche campo nome
        campo_nome = wait_for_element(*FORM_FIELDS["nome"],browser)
        campo_nome.send_keys(linha_dados["nome"])
        time.sleep(1)


        # Preenche campo email
        campo_email = wait_for_element(*FORM_FIELDS["email"],browser)
        campo_email.send_keys(linha_dados['email'])
        time.sleep(1)


        # Preenche campo telefone
        campo_telefone = wait_for_element(*FORM_FIELDS['telefone'], browser)
        campo_telefone.send_keys(linha_dados['telefone'])
        time.sleep(1)

        # Seleciona gênero (radio button)
        if linha_dados['genero'].lower() == 'masculino':
            radio = wait_for_element(*FORM_FIELDS['genero_masculino'],browser)
        else:
            radio = wait_for_element(*FORM_FIELDS['genero_feminino'],browser)

        radio.click()
        time.sleep(1)

        # Preenche campo sobre
        campo_sobre = wait_for_element(*FORM_FIELDS['sobre'],browser)
        campo_sobre.send_keys(linha_dados['sobre'])
        time.sleep(1)


        # Envia o formulário
        botao_enviar = wait_for_element(*FORM_FIELDS["submit"],browser)
        botao_enviar.click()
        time.sleep(2)


        print(f"✓ Formulário enviado para {linha_dados['nome']}")
        return True
    except TimeoutError as e:
        print(f"✗ Timeout ao processar {linha_dados.get('nome', 'N/A')}: {e}")
        return False
    except Exception as e:
        print(f"✗ Erro ao processar {linha_dados.get('nome', 'N/A')}: {e}")
        return False
    finally:
        if browser:
            browser.quit()
        






def carregar_dados_da_planilha(file: Path) -> list[dict]:
    """
    Carrega os dados da planilha Excel e retorna como lista de dicionários.

    Args:
        file (Path): Caminho do arquivo Excel contendo os dados.

    Returns:
        list[dict]: Lista de dicionários, cada um representando uma linha da planilha.
    """
    wb = load_workbook(file)
    ws = wb['Dados']


    dados = []

    for linha in range(2, ws.max_row + 1):
        linha_dados = {
            'nome': ws[f'A{linha}'].value or "",
            "email": ws[f'B{linha}'].value or "",
            "telefone": ws[f'C{linha}'].value or "",
            "genero": ws[f"D{linha}"].value or "",
            "sobre": ws[f'E{linha}'].value or "",
        }

        if not linha_dados["nome"]:
            continue

        dados.append(linha_dados)
    return dados 


if __name__ == "__main__":
    print("=== Iniciando preenchimento automático de formulários ===\n")


    if not FILE.exists():
        print(f"Erro: Arquivo {FILE} não encontrado.")
        exit(1)


    # Carrega dados da planilha.
    dados = carregar_dados_da_planilha(FILE)
    total = len(dados)


    if total == 0:
        print("Nenhum dado encontrado na planilha.")
        exit(0)

    print(f"Total de registros a processar: {total}\n")


    sucesso = 0
    falhas = 0

    for i, linha_dados  in enumerate(dados, 1):
        print(f"[{i}/{total}] ", end="")
        if preencher_planilha(linha_dados ):
            sucesso += 1
        else:
            falhas += 1
            print()

      # Resumo final
    print("=== Processamento concluído ===")
    print(f"Sucessos: {sucesso}")
    print(f"Falhas: {falhas}")
        












