"""
Módulo para consulta de CEPs via API ViaCEP e armazenamento
dos endereços em uma planilha Excel.

Fluxo:
    1. Lê CEPs de um arquivo Excel via 'return_ceps_in_file'.
    2. Consulta cada CEP na API ViaCEP via 'get_address_cep'.
    3. Salva os endereços encontrados em outro arquivo via 'save_address_in_excel'.
"""

import http.client
import json
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook

DIR = Path(__file__).parent
FILE_ADDRESS = DIR / "Address.xlsx"
FILE_CEPS = DIR / "CEPS.xlsx"

COLUMNS = {
    "CEP":        "cep",
    "Logradouro": "logradouro",
    "Bairro":     "bairro",
    "Localidade": "localidade",
}


def get_address_cep(cep: str) -> dict[str, str]:
    """
    Consulta a API ViaCEP e retorna os dados de endereço do CEP informado.

    Args:
        cep (str): CEP a ser consultado, somente números (ex: '01001000').

    Returns:
        dict[str, str]: Dicionário com os dados do endereço se encontrado.
        dict vazio: Se o CEP for inválido ou ocorrer erro na requisição.
    """
    connection = http.client.HTTPSConnection("viacep.com.br")
    try:
        # Envia requisição GET para a API ViaCEP com o CEP informado
        connection.request("GET", f"/ws/{cep}/json")
        response = connection.getresponse()

        # Decodifica a resposta JSON em UTF-8
        address = json.loads(response.read().decode("utf-8"))

        # A API retorna {"erro": true} quando o CEP não existe
        if "erro" in address:
            print(f"CEP {cep} não encontrado.")
            return {}

        return address

    except Exception as e:
        print(f"Erro na requisição do CEP {cep}: {e}")
        return {}

    finally:
        # Garante o fechamento da conexão independente do resultado
        connection.close()


def return_ceps_in_file(file: str) -> list[str]:
    """
    Lê e retorna os CEPs armazenados na primeira coluna de um arquivo Excel.

    Args:
        file (str): Caminho do arquivo Excel contendo os CEPs.

    Returns:
        list[str]: Lista de CEPs lidos a partir da segunda linha (ignora cabeçalho).
    """
    wb = load_workbook(file)
    ws = wb.active
    ceps = []

    # Itera a partir da linha 2 para ignorar o cabeçalho
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                # Remove hífen para padronizar o formato do CEP
                ceps.append(str(cell.value).replace("-", ""))

    return ceps


def save_address_in_excel(addresses: list[dict[str, str]], file: str) -> None:
    """
    Salva uma lista de endereços em um arquivo Excel, evitando duplicatas.

    Cria o arquivo com cabeçalho se não existir. Caso já exista,
    carrega os dados e adiciona apenas os CEPs ainda não salvos.

    Args:
        addresses (list[dict[str, str]]): Lista de endereços retornados pela API ViaCEP.
        file (str): Caminho do arquivo Excel de destino.
    """
    # Cria arquivo novo com cabeçalho ou carrega o existente
    if os.path.exists(file):
        wb = load_workbook(file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(list(COLUMNS.keys()))

    # Coleta CEPs já salvos para evitar duplicatas
    saved_ceps = {
        ws.cell(row=i, column=1).value 
        for i in range(2, ws.max_row + 1)
    }

    for address in addresses:
        cep = address.get("cep")

        if cep in saved_ceps:
            print(f"CEP {cep} já salvo, pulando...")
            continue

        # Salva apenas as colunas definidas em COLUMNS
        ws.append([address.get(key, "N/A") for key in COLUMNS.values()])
        saved_ceps.add(cep)
        print(f"CEP {cep} salvo.")

    wb.save(file)


def add_ceps_in_file(file: str, ceps: list[str]) -> None:
    """
    Cria um arquivo Excel com uma lista de CEPs na coluna 'A'.
    Utilizado para popular o arquivo de entrada para testes.

    Args:
        file (str): Caminho do arquivo Excel a ser criado.
        ceps (list[str]): Lista de CEPs a serem inseridos.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "CEPS"

    # Insere cada CEP a partir da segunda linha
    for i, cep in enumerate(ceps, start=2):
        ws[f"A{i}"] = cep

    wb.save(file)


if __name__ == "__main__":
    ceps_test = ["01001000", "77001-432", "79117-440"]

    # Cria arquivo de CEPs para teste
    add_ceps_in_file(FILE_CEPS, ceps_test)

    # Lê, consulta e salva os endereços
    addresses = []
    for cep in return_ceps_in_file(FILE_CEPS):
        result = get_address_cep(cep)
        if result:
            addresses.append(result)

    save_address_in_excel(addresses, FILE_ADDRESS)